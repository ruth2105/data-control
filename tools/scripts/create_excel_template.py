import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Reagent Consumption"

# Reagent names (32 products)
reagents = ['TSH','T3','T4','FT3','FT4','Vitamin D','CK-MB','D-Dimer','Ferritin',
    'Vitamin B12','Folate','AMH','Estradiol','FSH','LH','Prolactin','Progesterone',
    'Testosterone','AFP','CEA','CA19-9','CA125','β-HCG','HSCTNL','cTnl','PCT','TPSA',
    'fPSA','Anti-HIV','HBsAg','Anti-HCV','Anti-TP']

# Customer names (64 total)
customers = [
    "Lucy health care plc","AL-FEWZ Health care","Dr.Meliha Gyn/Obs",
    "Lubu Internal and MCH Clinic","Semah MCH","UNECA","Care Medium clinic","Nuri Enat MCH",
    "Bati Hospital","RiftVally Hospital","Tirat Rad PLC","Geda Medical Laboratory",
    "Dr. Ermias Medium clinic","Jimma Awetu primary Hospital","Abajifar Medium Clinic (A90)",
    "Abajifar Medium Clinic (A8)","Dr.Belete Medium Clinic","Shenen Gibe General Hospital",
    "My Medium Clinic","Dr. Tesfa (Kidus Medium Clinic)","Vital Internal Medicine",
    "Arada Georgis Specialized Medical center","Biruhkids Pediatrics center",
    "Dr.Shemse MCH Center","Fikir Abrak internal and MCH center (A8)","Lobe Midium Clinic",
    "Legehar General Hospital","Senay Higher Clinic","Fikir Abrak internal and MCH center (A90)",
    "Qelebet specialised clinic","Uropshare Medical Service","Summit Health Center",
    "Universe Internal Medicine Speciality Clinic","Hiwot Fana Hospital","Hamdi Specialized Clinic",
    "Robsen Clinic","Odo dola Midium Clinic","Dilla University referral Hospital",
    "Kibru primary hospital","Unicare medical service","Newyork Health care",
    "Cardinal Speciality Medical & Surgical Center","Grace Hospital",
    "Guro Basic Diagnostic Laboratory","Arba Minch General Hospital","Hayat Hospital",
    "Amin General Hospital","Caring Internal Medicine Speciality Clinic","Addis Cardiac Hospital",
    "Makira Hakime Pediatric and Surgical Center","Guro Basic Diagnostic Laboratory (No 02)",
    "Birhan Ethiopia international Medicine","Alem Tena Internal Medicine Speciality Center",
    "ACE Engineeering PLC","Alta Medium Clinic","Medhen Beza Primary Hospital",
    "Alem Gena Medical Services","American Medical Center","Alatyon General Hospital",
    "Feya General Hospital","Alem Primary Hospital","Rohobot General Hospital",
    "Mohammed Akle General Hospital","Sold not installed"
]

# Define styles
green_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
light_gray_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
bold_font = Font(name='Calibri', size=11, bold=True)
normal_font = Font(name='Calibri', size=11)
thick_border = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='medium', color='000000')
)
double_top_border = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='double', color='000000'),
    bottom=Side(style='medium', color='000000')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# Set column widths
ws.column_dimensions['A'].width = 45  # Customer name
for i in range(2, 34):  # Reagent columns
    ws.column_dimensions[get_column_letter(i)].width = 13
ws.column_dimensions[get_column_letter(34)].width = 12  # Total
ws.column_dimensions[get_column_letter(35)].width = 10  # %

# Header row
ws['A1'] = 'Customer Name'
ws['A1'].fill = green_fill
ws['A1'].font = white_font
ws['A1'].alignment = center_align
ws['A1'].border = thick_border

for idx, reagent in enumerate(reagents, start=2):
    cell = ws.cell(row=1, column=idx)
    cell.value = reagent
    cell.fill = green_fill
    cell.font = white_font
    cell.alignment = center_align
    cell.border = thick_border

# Total and % headers
ws.cell(row=1, column=34).value = 'Total'
ws.cell(row=1, column=34).fill = green_fill
ws.cell(row=1, column=34).font = white_font
ws.cell(row=1, column=34).alignment = center_align
ws.cell(row=1, column=34).border = thick_border

ws.cell(row=1, column=35).value = '%'
ws.cell(row=1, column=35).fill = green_fill
ws.cell(row=1, column=35).font = white_font
ws.cell(row=1, column=35).alignment = center_align
ws.cell(row=1, column=35).border = thick_border

# Customer rows (data cells)
for row_idx, customer in enumerate(customers, start=2):
    # Customer name
    cell = ws.cell(row=row_idx, column=1)
    cell.value = customer
    cell.font = bold_font
    cell.alignment = left_align
    cell.border = thick_border
    
    # Reagent columns (empty for user to fill)
    for col_idx in range(2, 34):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = 0
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thick_border
    
    # Total column with formula
    total_cell = ws.cell(row=row_idx, column=34)
    total_cell.value = f"=SUM(B{row_idx}:AF{row_idx})"
    total_cell.fill = light_gray_fill
    total_cell.font = bold_font
    total_cell.alignment = center_align
    total_cell.border = thick_border
    
    # % column with formula
    percent_cell = ws.cell(row=row_idx, column=35)
    percent_cell.value = f"=IF($AH${len(customers)+2}=0,0,AH{row_idx}/$AH${len(customers)+2})"
    percent_cell.number_format = '0.0%'
    percent_cell.fill = light_gray_fill
    percent_cell.font = bold_font
    percent_cell.alignment = center_align
    percent_cell.border = thick_border

# TOTAL row
total_row = len(customers) + 2
ws.cell(row=total_row, column=1).value = 'TOTAL'
ws.cell(row=total_row, column=1).fill = gray_fill
ws.cell(row=total_row, column=1).font = bold_font
ws.cell(row=total_row, column=1).alignment = center_align
ws.cell(row=total_row, column=1).border = double_top_border

# Total formulas for each reagent
for col_idx in range(2, 34):
    cell = ws.cell(row=total_row, column=col_idx)
    col_letter = get_column_letter(col_idx)
    cell.value = f"=SUM({col_letter}2:{col_letter}{len(customers)+1})"
    cell.fill = gray_fill
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = double_top_border

# Grand total
grand_total = ws.cell(row=total_row, column=34)
grand_total.value = f"=SUM(AH2:AH{len(customers)+1})"
grand_total.fill = gray_fill
grand_total.font = bold_font
grand_total.alignment = center_align
grand_total.border = double_top_border

# 100% cell
percent_total = ws.cell(row=total_row, column=35)
percent_total.value = "100%"
percent_total.fill = gray_fill
percent_total.font = bold_font
percent_total.alignment = center_align
percent_total.border = double_top_border

# Set row heights
for row in range(1, total_row + 1):
    ws.row_dimensions[row].height = 20

# Freeze panes (freeze header row and first column)
ws.freeze_panes = 'B2'

# Save file
wb.save('Quarterly_Reagent_Consumption_Template.xlsx')
print("✅ Excel template created: Quarterly_Reagent_Consumption_Template.xlsx")
print("📊 64 customers × 32 reagents with full formatting, borders, and colors!")
